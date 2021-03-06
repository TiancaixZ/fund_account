from openpyxl import load_workbook
from data import Summary_data, Inc_Exp_data
from util import name_list


class Read:

    def __init__(self, password, address, date, show):
        self._password = password
        self._address = address
        self._date = date
        self._show = show
        self.wb = load_workbook(self._address)

    def read_workbook(self):
        wb = load_workbook(self._address)
        return wb

    @property
    def summary_list(self):
        summary_list = []
        for sheet in self.wb:
            total_income = 0
            total_expenditure = 0
            total_contacts = 0
            total_lmb = 0

            for row in range(4, sheet.max_row + 1):
                if self.lmb(row, sheet) is not None:
                    total_lmb = self.lmb(row, sheet)
                total_income = total_income + self.income_expenditure(7, row, sheet)
                total_expenditure = total_expenditure + self.income_expenditure(8, row, sheet)
                total_contacts = total_contacts + self.contacts(row, sheet)

            summary_data = Summary_data(sheet.title, total_lmb, total_income,
                                        total_expenditure, total_contacts)
            summary_list.append(summary_data)
            self._show.show("成功读取" + sheet.title + "汇总" + "..............................OK")
            print("成功读取" + sheet.title + "汇总" + "..............................OK")
        return summary_list

    @property
    def pri_inc_list(self):
        return self.pri_com_exp(0)

    @property
    def pri_exp_list(self):
        return self.pri_com_exp(1)

    @property
    def com_inc_list(self):
        return self.pri_com_exp(2)

    @property
    def com_exp_list(self):
        return self.pri_com_exp(3)

    def pri_com_exp(self, pri_com_inc_exp):
        """
        负责人/公司 收入/支出
        :param pri_com_inc_exp:
        :return:
        """

        data_list = []

        col = 0
        col_1 = 0
        if pri_com_inc_exp == 0:
            col = 10
            col_1 = 7  # 负责人 收入
            print_str = "负责人 收入"
        elif pri_com_inc_exp == 1:
            col = 10
            col_1 = 8  # 负责人 支出
            print_str = "负责人 支出"
        elif pri_com_inc_exp == 2:
            col = 6
            col_1 = 7  # 公司 收入
            print_str = "公司 收入"
        elif pri_com_inc_exp == 3:
            col = 6
            col_1 = 8  # 公司 支出
            print_str = "公司 支出"

        for sheet in self.wb:
            if sheet.title != "理财":
                if pri_com_inc_exp == 2 or pri_com_inc_exp == 3:
                    if sheet.title == "民生对公" or sheet.title == "承兑汇票" or sheet.title == "半额承兑":
                        com_name_list = name_list(1, sheet)
                        val = 0
                        for name in com_name_list:
                            for row in range(4, sheet.max_row + 1):
                                if sheet.cell(row, col).value == name and \
                                        sheet.cell(row, col_1).value is not None and \
                                        sheet.cell(row, 3).value != "往来" and \
                                        sheet.cell(row, 1).value == int(self._date):
                                    val = sheet.cell(row, col_1).value + val
                            data = Inc_Exp_data(sheet.title, name, val)
                            data_list.append(data)
                            val = 0
                        self._show.show("成功读取" + sheet.title + print_str + "..............................OK")
                        print("成功读取" + sheet.title + print_str + "..............................OK")
                else:
                    pri_name_list = name_list(0, sheet)
                    val = 0
                    for name in pri_name_list:
                        for row in range(4, sheet.max_row + 1):
                            if sheet.cell(row, col).value == name and \
                                    sheet.cell(row, col_1).value is not None and \
                                    sheet.cell(row, 3).value != "往来" and \
                                    sheet.cell(row, 1).value == int(self._date):
                                val = sheet.cell(row, col_1).value + val
                        data = Inc_Exp_data(sheet.title, name, val)
                        data_list.append(data)
                        val = 0
                    self._show.show("成功读取" + sheet.title + print_str + "..............................OK")
                    print("成功读取" + sheet.title + print_str + "..............................OK")
        return data_list

    def lmb(self, row, sheet):
        """
        上月余额
        :param row:
        :param sheet:
        :return:
        """
        month = str(self._date)[2:4]
        if month == "01":
            if sheet.cell(row, 3).value == "上年结转" and \
                    sheet.cell(row, 1).value == int(self._date):
                if sheet.cell(row, 9).value is not None:
                    return sheet.cell(row, 9).value
                else:
                    return -1
        else:
            if sheet.cell(row, 3).value == "上月结转" and \
                    sheet.cell(row, 1).value == int(self._date):
                if sheet.cell(row, 9).value is not None:
                    return sheet.cell(row, 9).value
                else:
                    return -1

    def income_expenditure(self, types, row, sheet):
        """
        收入 支出
        :param types:
        :param row:
        :param sheet:
        :return:
        """
        val = 0.00

        """往来 上年结转 上月结转这三个特殊的摘要 其余为收入 支出"""
        if sheet.cell(row, 3).value != "往来" and \
                sheet.cell(row, 3).value != "上年结转" and \
                sheet.cell(row, 3).value != "上月结转" and \
                sheet.cell(row, 1).value == int(self._date):
            if sheet.cell(row, types).value is not None:
                val = sheet.cell(row, types).value

        return val

    def contacts(self, row, sheet):
        """
        往来
        :param row:
        :param sheet:
        :return:
        """
        val = 0
        if sheet.cell(row, 3).value == "往来" and \
                sheet.cell(row, 1).value == int(self._date):
            if sheet.cell(row, 7).value is not None:
                val = sheet.cell(row, 7).value
            else:
                val = -sheet.cell(row, 8).value

        return val

# def read_worksheet(password, address):
#     """
#     解密excel
#     :param password:
#     :param address:
#     :return:
#     """
#     xlapp = win32com.client.Dispath("excel.Application")
#     wb = xlapp.Workbooks.open(address, False, True, None, password)
#
#     if wb is not None:
#         print("test")
#     else:
#         print("no")
#
#     return wb
#
#
# def read_list(date, address, show):
#     """
#     获取所有列表
#     :param show:
#     :param date:
#     :param address:
#     :return:
#     """
#     summary_list = []
#
#     wb = load_workbook(address)
#     for sheet in wb:
#         total_income = 0
#         total_expenditure = 0
#         total_contacts = 0
#         total_lmb = 0
#
#         for row in range(4, sheet.max_row + 1):
#             if lmb(date, row, sheet) is not None:
#                 total_lmb = lmb(date, row, sheet)
#             total_income = total_income + income_expenditure(7, date, row, sheet)
#             total_expenditure = total_expenditure + income_expenditure(8, date, row, sheet)
#             total_contacts = total_contacts + contacts(date, row, sheet)
#
#         summary_data = Summary_data(sheet.title, total_lmb, total_income,
#                                     total_expenditure, total_contacts)
#         summary_list.append(summary_data)
#         show.show("成功读取" + sheet.title + "汇总" + "..............................OK")
#         print("成功读取" + sheet.title + "汇总" + "..............................OK")
#
#     pri_inc_list = pri_com_exp(date, 0, wb)
#     pri_exp_list = pri_com_exp(date, 1, wb)
#     com_inc_list = pri_com_exp(date, 2, wb)
#     com_exp_list = pri_com_exp(date, 3, wb)
#     return summary_list, pri_inc_list, pri_exp_list, com_inc_list, com_exp_list
#
#
# def lmb(date, row, sheet):
#     """
#     上月余额
#     :param date:
#     :param row:
#     :param sheet:
#     :return:
#     """
#     month = str(date)[2:4]
#     if month == "01":
#         if sheet.cell(row, 3).value == "上年结转" and \
#                 sheet.cell(row, 1).value == int(date):
#             if sheet.cell(row, 9).value is not None:
#                 return sheet.cell(row, 9).value
#             else:
#                 return -1
#     else:
#         if sheet.cell(row, 3).value == "上月结转" and \
#                 sheet.cell(row, 1).value == int(date):
#             if sheet.cell(row, 9).value is not None:
#                 return sheet.cell(row, 9).value
#             else:
#                 return -1
#
#
# def income_expenditure(types, date, row, sheet):
#     """
#     收入 支出
#     :param types:
#     :param date:
#     :param row:
#     :param sheet:
#     :return:
#     """
#     val = 0.00
#
#     """往来 上年结转 上月结转这三个特殊的摘要 其余为收入 支出"""
#     if sheet.cell(row, 3).value != "往来" and \
#             sheet.cell(row, 3).value != "上年结转" and \
#             sheet.cell(row, 3).value != "上月结转" and \
#             sheet.cell(row, 1).value == int(date):
#         if sheet.cell(row, types).value is not None:
#             val = sheet.cell(row, types).value
#
#     return val
#
#
# def contacts(date, row, sheet):
#     """
#     往来
#     :param date:
#     :param row:
#     :param sheet:
#     :return:
#     """
#     val = 0
#     if sheet.cell(row, 3).value == "往来" and \
#             sheet.cell(row, 1).value == int(date):
#         if sheet.cell(row, 7).value is not None:
#             val = sheet.cell(row, 7).value
#         else:
#             val = -sheet.cell(row, 8).value
#
#     return val
#
#
# def pri_com_exp(date, pri_com_inc_exp, wb):
#     """
#     负责人/公司 收入/支出
#     :param date:
#     :param pri_com_inc_exp:
#     :param wb:
#     :return:
#     """
#
#     data_list = []
#
#     col = 0
#     col_1 = 0
#     if pri_com_inc_exp == 0:
#         col = 10
#         col_1 = 7  # 负责人 收入
#         print_str = "负责人 收入"
#     elif pri_com_inc_exp == 1:
#         col = 10
#         col_1 = 8  # 负责人 支出
#         print_str = "负责人 支出"
#     elif pri_com_inc_exp == 2:
#         col = 6
#         col_1 = 7  # 公司 收入
#         print_str = "公司 收入"
#     elif pri_com_inc_exp == 3:
#         col = 6
#         col_1 = 8  # 公司 支出
#         print_str = "公司 支出"
#
#     for sheet in wb:
#         if sheet.title != "理财":
#             if pri_com_inc_exp == 2 or pri_com_inc_exp == 3:
#                 if sheet.title == "民生对公" or sheet.title == "承兑汇票" or sheet.title == "半额承兑":
#                     com_name_list = name_list(1, sheet)
#                     val = 0
#                     for name in com_name_list:
#                         for row in range(4, sheet.max_row + 1):
#                             if sheet.cell(row, col).value == name and \
#                                     sheet.cell(row, col_1).value is not None and \
#                                     sheet.cell(row, 3).value != "往来" and \
#                                     sheet.cell(row, 1).value == int(date):
#                                 val = sheet.cell(row, col_1).value + val
#                         data = Inc_Exp_data(sheet.title, name, val)
#                         data_list.append(data)
#                         val = 0
#                     print("成功读取" + sheet.title + print_str + "..............................OK")
#             else:
#                 pri_name_list = name_list(0, sheet)
#                 val = 0
#                 for name in pri_name_list:
#                     for row in range(4, sheet.max_row + 1):
#                         if sheet.cell(row, col).value == name and \
#                                 sheet.cell(row, col_1).value is not None and \
#                                 sheet.cell(row, 3).value != "往来" and \
#                                 sheet.cell(row, 1).value == int(date):
#                             val = sheet.cell(row, col_1).value + val
#                     data = Inc_Exp_data(sheet.title, name, val)
#                     data_list.append(data)
#                     val = 0
#                 print("成功读取" + sheet.title + print_str + "..............................OK")
#     return data_list
#
#
# def name_list(pri_inc_row, sheet):
#     """
#     单位/负责人  去重
#     :param pri_inc_row:
#     :param sheet:
#     :return:
#     """
#     old_list = []
#     if pri_inc_row == 0:
#         for row in range(4, sheet.max_row + 1):
#             if sheet.cell(row, 10).value is not None:
#                 old_list.append(sheet.cell(row, 10).value)  # 负责人
#     else:
#         for row in range(4, sheet.max_row + 1):
#             if sheet.cell(row, 6).value is not None:
#                 old_list.append(sheet.cell(row, 6).value)  # 单位名称
#     new_list = list(set(old_list))
#
#     return new_list
