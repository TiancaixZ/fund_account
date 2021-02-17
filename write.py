from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment

from util import if_null, muban_list


class Write:
    
    def __init__(self, date, path, show):
        self._date = date
        self._path = path
        self._show = show

        """创建sheet"""
        self.write_wb = Workbook()
        self.ws_summary = self.write_wb.create_sheet("汇总", 0)
        self.ws_income_principal = self.write_wb.create_sheet("收入_负责人", 1)
        self.ws_expenditure_principal = self.write_wb.create_sheet("支出_负责人", 2)
        self.ws_income_company = self.write_wb.create_sheet("收入_公司", 3)
        self.ws_expenditure_company = self.write_wb.create_sheet("支出_公司", 4)

    def summary_sheet(self, summary_list):
        """
        汇总
        :return:
        """
        self.ws_summary["A1"] = "2021年" + str(self._date)[2:4] + "月份朗坤资金流向"
        self.ws_summary["A1"].border = sheet_border()
        self.ws_summary["A1"].alignment = sheet_alignment()
        self.ws_summary["A2"] = "科目"
        self.ws_summary["B2"] = "上月余额"
        self.ws_summary["C2"] = "收入"
        self.ws_summary["D2"] = "支出"
        self.ws_summary["E2"] = "往来"
        self.ws_summary["F2"] = "余额"
        self.ws_summary.merge_cells('A1:F1')
        for col in self.ws_summary['A2:F2']:
            for cell in col:
                cell.border = sheet_border()
                cell.alignment = sheet_alignment()
        self.ws_summary.column_dimensions['A'].width = 20  # 设置列宽20
        self._show.show("汇总 表头创建..............................OK")
        print("汇总 表头创建..............................OK")

        total_lmb = 0
        total_expenditure = 0
        total_income = 0
        total_contacts = 0
        total_balance = 0
        for row in range(3, len(summary_list) + 4):
            if row == len(summary_list) + 3:
                _ = self.ws_summary.cell(row, 1, "合计")
                _ = self.ws_summary.cell(row, 2, total_lmb)
                _ = self.ws_summary.cell(row, 3, total_income)
                _ = self.ws_summary.cell(row, 4, total_expenditure)
                _ = self.ws_summary.cell(row, 5, total_contacts)
                _ = self.ws_summary.cell(row, 6, total_balance)
            else:
                _ = self.ws_summary.cell(row, 1, summary_list[row - 3].name)
                _ = self.ws_summary.cell(row, 2, summary_list[row - 3].lmb)
                total_lmb = total_lmb + if_null(summary_list[row - 3].lmb)
                _ = self.ws_summary.cell(row, 3, summary_list[row - 3].income)
                total_income = total_income + summary_list[row - 3].income
                _ = self.ws_summary.cell(row, 4, summary_list[row - 3].expenditure)
                total_expenditure = total_expenditure + summary_list[row - 3].expenditure
                _ = self.ws_summary.cell(row, 5, summary_list[row - 3].contacts)
                total_contacts = total_contacts + summary_list[row - 3].contacts
                _ = self.ws_summary.cell(row, 6, summary_list[row - 3].balance)
                total_balance = total_balance + summary_list[row - 3].balance
            for col in range(1, 7):
                _ = self.ws_summary.cell(row, col).border = sheet_border()
                _ = self.ws_summary.cell(row, col).alignment = sheet_alignment()
        self._show.show("汇总 数据填写..............................OK")
        print("汇总 数据填写..............................OK")

    def muban(self, name_list, pricom_list, ws):
        """
        模板生成 表头 首列
        :param name_list:
        :param pricom_list:
        :param ws:
        :return:
        """
        ws["A1"].border = sheet_border()
        ws["A1"].alignment = sheet_alignment()
        for column in range(2, len(name_list)+3):
            if column == len(name_list)+2:
                pass
            else:
                _ = ws.cell(column=column, row=2, value=name_list[column-2])  # 卡名
        for row in range(3, len(pricom_list) + 4):
            if row == len(pricom_list) + 3:
                _ = ws.cell(column=1, row=row, value="合计")
            else:
                _ = ws.cell(column=1, row=row, value=pricom_list[row-3])  # 单位名称/负责人

        for col in range(1, len(name_list) + 2):
            for row in range(2, len(pricom_list) + 5):
                ws.cell(row, col).border = sheet_border()
                ws.cell(row, col).alignment = sheet_alignment()
        ws.column_dimensions['A'].width = 40  # 设置列宽40
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(name_list))  # 合并单元表头
        self._show.show("模板创建..............................OK")
        print("模板创建..............................OK")

    def incom_exp_sheet(self, ws, month, ready_list, pri_com_inc_exp):
        """
        公司/负责人 收入/支出 表
        :param ws:
        :param month:
        :param ready_list:
        :param pri_com_inc_exp:
        :return:
        """
        if pri_com_inc_exp == 0:
            var = "负责人"
            var_title = "收入"
        elif pri_com_inc_exp == 1:
            var = "负责人"
            var_title = "支出"
        elif pri_com_inc_exp == 2:
            var = "单位名称"
            var_title = "收入"
        elif pri_com_inc_exp == 3:
            var = "单位名称"
            var_title = "支出 "

        ws["A1"] = "202年" + month + "月份朗坤资金" + var_title
        ws["A2"] = var
        name_list, pricom_list = muban_list(ready_list)
        print(len(name_list))
        self.muban(name_list, pricom_list, ws)

        for item in ready_list:
            for row in range(3, len(pricom_list) + 4):
                for column in range(2, len(name_list) + 3):
                    if row == len(pricom_list) + 3:
                        pass
                    elif column == len(name_list) + 2:
                        pass
                    elif item.name == name_list[column-2] and item.pri_com == pricom_list[row-3]:
                        _ = ws.cell(column=column, row=row, value=item.inc_exp)
    
    def ws_income_principal_sheet(self, pri_inc_list):
        self.incom_exp_sheet(self.ws_income_principal, str(self._date)[2:4], pri_inc_list, 0)

    def ws_expenditure_principal_sheet(self, pri_exp_list):
        self.incom_exp_sheet(self.ws_expenditure_principal, str(self._date)[2:4], pri_exp_list, 1)

    def ws_income_company_sheet(self, com_inc_list):
        self.incom_exp_sheet(self.ws_income_company, str(self._date)[2:4], com_inc_list, 2)

    def ws_expenditure_company_sheet(self, com_exp_list):
        self.incom_exp_sheet(self.ws_expenditure_company, str(self._date)[2:4], com_exp_list, 3)

    def save(self):
        path = self._path + "/" + "test.xlsx"
        self.write_wb.save(path)

        self._show.show("文件保存成功..............................OK")
        print("文件保存成功..............................OK")


def sheet_border():
    """
    边框
    :return:
    """
    thin = Side(border_style="thin", color="000000")  # 边框样式，颜色
    border = Border(left=thin, right=thin, top=thin, bottom=thin)  # 边框的位置

    return border


def sheet_alignment():
    """
    对齐
    :return:
    """
    align = Alignment(horizontal='center', vertical='center')  # 剧中对齐
    return align
